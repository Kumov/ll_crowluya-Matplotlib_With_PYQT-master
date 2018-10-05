# -*- coding: utf-8 -*-
'''
    【简介】
	PyQt5中 QFileDialog 例子
   
  
'''
import sys
import openpyxl
import pandas as pd  # 引入pandas
import os

from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *


log = print


class filedialogdemo(QWidget):
    def __init__(self, parent=None):
        super(filedialogdemo, self).__init__(parent)
        layout = QVBoxLayout()

        self.btn1 = QPushButton("加载文本文件")
        self.btn1.clicked.connect(self.getfiles)
        layout.addWidget(self.btn1)

        self.process_btn = QPushButton("处理文件")
        self.process_btn.clicked.connect(self.process_file)
        layout.addWidget(self.process_btn)

        self.save_btn = QPushButton("保存txt")
        self.save_btn.clicked.connect(self.save_txt_file)
        layout.addWidget(self.save_btn)

        self.save_excle_btn = QPushButton("保存excel")
        self.save_excle_btn.clicked.connect(self.save_excel_file)
        layout.addWidget(self.save_excle_btn)

        self.le = QLabel("")
        self.contents = QTextEdit()
        layout.addWidget(self.contents)
        layout.addWidget(self.le)

        self.setLayout(layout)
        self.setWindowTitle("File Dialog 例子")

    def getfile(self):
        fname, _ = QFileDialog.getOpenFileName(self, 'Open file', 'c:\\', "All Files (*);;Text Files (*.xlsx)")
        self.le.setPixmap(QPixmap(fname))

    def save_txt_file(self):
        fname, _ = QFileDialog.getSaveFileName(self,
                                                     "文件保存",
                                                     "C:/",
                                                     "Text Files (*.txt)")

        self.le.setPixmap(QPixmap(fname))

        # 获取保存的文件夹地址
        [dirname, filename] = os.path.split(fname)
        print(dirname, "\n", filename)
        self.save_path_name = dirname
        self.xlsx_to_txt()
        log('debug 保存文件({})'.format(fname))

    def save_excel_file(self):
        fname, _ = QFileDialog.getSaveFileName(self,
                                                     "文件保存",
                                                     "C:/",
                                                     "Text Files (*.xlsx)")

        self.le.setPixmap(QPixmap(fname))
        # 获取保存的文件夹地址
        [dirname, filename] = os.path.split(fname)
        log('debug save_excel_file 保存文件({})'.format(dirname))
        print(dirname, "\n", filename)
        # 保存文件到excel
        self.save_path_name_excel = dirname
        self.xlsx_to_excel()
        log('debug 保存文件({})'.format(fname))

    def getfiles(self):
        dlg = QFileDialog()
        dlg.setFileMode(QFileDialog.AnyFile)
        dlg.setFilter(QDir.Files)

        if dlg.exec_():
            filenames = dlg.selectedFiles()
            f = open(filenames[0], 'r')
            self.file_loaded_name = filenames[0]
            log("加载的文件filename({})".format(filenames[0]))
            with f:
                data = f.read()
                self.contents.setText(data)
                log("加载完成")
            return filenames[0]

    def process_file(self):
        log(" 点击处理文件")
        path = self.file_loaded_name
        self.txt_to_excel(path)
        self.msg_process_success()

        # data = self.getfiles()
        # log(type(data))
        # # data_list = data.split('\r\n')
        # log(type(data), data)
        pass

    def txt_to_excel(self, fpath):
        filename = "微相.txt"
        filepath = fpath
        log("保存文件的文件名字({})".format(filepath))

        filename = filepath.split('/')[-1]
        papa = pd.read_csv(filepath_or_buffer=filepath, sep='\t', encoding="GBK")  # 加载papa.txt,指定它的分隔符是 \t
        # papa.set_index('井名')
        # log("打开文件({})", papa)
        self.pdates = papa
        writer = pd.ExcelWriter('{}.xlsx'.format(filename))
        self.xlsx_processed_name = '{}.xlsx'.format(filename)
        papa.to_excel(writer, 'sheet1')
        self.xlsx_sheet = 'sheet1'
        log("到ec 保存")
        writer.save()
        log("end 保存")

    def xlsx_to_txt(self, papa=""):
        # papa是pandas打开的文件
        log("开始保存处理过的数据")
        papa = self.pdates

        filename = self.xlsx_processed_name
        log("filename", self.file_loaded_name)
        workbook = openpyxl.load_workbook(filename)
        log("创建了work")
        # 所有的工作表单名字
        # log(workbook.sheetnames)

        # 拿到某个 工作表 内容的标准做法，我们这个 excel 文件只有一个三条曲线数据工作表
        sheet = workbook['sheet1']

        maxrow = sheet.max_row  # 最大行数
        log('maxrow', maxrow)

        log('row', sheet['B'])
        # # 遍历第二行所有格子
        # for c in sheet['2']:
        #     log('格子({})'.format(c.value))
        # # 遍历获取 A 列所有的值
        # log('start time')

        # log('start time')
        wells = []
        wels = set()
        for a in sheet['B']:
            # log('a', a.value)
            if a.value == "井名":
                head = a.value
            else:
                wells.append(a.value)
            # wels.add(a)
        # wels = set(wells)

        news_ids = []
        for id in wells:
            if id not in news_ids:
                news_ids.append(id)
        # wels.clear()
        log('end time', len(news_ids))
        # log('end time', news_ids)

        # 得到不重复的井名, 保存为excel
        path = self.save_path_name

        log("txt path", path)

        for w in news_ids:
            c = papa[papa['井名'].isin([w])]
            # wname = '{}井.xlsx'.format(w)
            wname = '{}井.txt'.format(w)
            log(wname)
            log(path)
            # dparh =
            dpath = path + '/' + wname
            # writer = pd.ExcelWriter(wname)

            log('# 输出成txt', dpath)

            c.to_csv(path_or_buf=dpath, sep=' ')
            log('#完成 输出成txt')

            # 输出成excel
            # c.to_excel(writer, 'Sheet1')
            # df2.to_excel(writer, 'Sheet2')
            # writer.save()
            log("保存txt结束 ", w)
            # break

        # 弹出保存数据成功
        self.msg_save_txt_success()

    def xlsx_to_excel(self, papa=""):
        # papa是pandas打开的文件
        log("开始保存处理过的数据 到 excel")
        papa = self.pdates

        filename = self.xlsx_processed_name
        log("filename", self.xlsx_processed_name)
        workbook = openpyxl.load_workbook(filename)
        log("创建了work")
        # 所有的工作表单名字
        # log(workbook.sheetnames)
        sheet = workbook['sheet1']

        maxrow = sheet.max_row  # 最大行数
        log('maxrow', maxrow)
        wells = []
        wels = set()
        for a in sheet['B']:
            # log('a', a.value)
            if a.value == "井名":
                head = a.value
            else:
                wells.append(a.value)

        # 井名去重
        news_ids = []
        for id in wells:
            if id not in news_ids:
                news_ids.append(id)
        log('当前共口{}井'.format(len(news_ids)))

        # 得到不重复的井名, 保存为excel
        # path为选中保存的dir
        path = self.save_path_name_excel
        # log('end time({}) ({})'.format(len(path), path))

        for w in news_ids:
            c = papa[papa['井名'].isin([w])]
            # wname = '{}井.xlsx'.format(w)
            wname = '{}井.xlsx'.format(w)
            log(wname)
            log(path)
            epath = path

            writer = pd.ExcelWriter(wname)
            log('# 输出成excel', epath)
            # 输出成excel
            # path = 'C:\'

            writer = pd.ExcelWriter('{}/{}.xlsx'.format(epath, wname)) # 指定Excel文件名与路径

            c.to_excel(writer, sheet_name='Sheet1')
            writer.save()
            # break
        log("保存excel结束 ", w)
        # 弹出保存成功窗口
        self.msg_save_excel_success()


    def msg_process_success(self, msg=''):
        # 弹出关于此软件的信息
        msg = """ 文件处理:

        操作处理成功
        """
        QMessageBox.about(self, "About the demo", msg.strip())

    def msg_save_txt_success(self, msg=''):
        # 弹出关于此软件的信息
        msg = """ 保存成txt文件:

        操作处理成功
        """
        QMessageBox.about(self, "About the demo", msg.strip())

    def msg_save_excel_success(self, msg=''):
        # 弹出关于此软件的信息
        msg = """ 保存成excel文件:

         操作处理成功
         """
        QMessageBox.about(self, "About the demo", msg.strip())

def test():
    # app = QApplication(sys.argv)
    # ex = filedialogdemo()
    # ex.process_file()

    txt_to_excel("微相.txt")
    log("测试完成")

# 1.加载excel数据
# 2.根据excel数据切割


def main():
    app = QApplication(sys.argv)
    ex = filedialogdemo()

    # ex.process_file()
    ex.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
    # test()
