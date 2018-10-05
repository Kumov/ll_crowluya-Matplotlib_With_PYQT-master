# -*- coding: utf-8 -*-
'''
    【简介】
	PyQt5中 QFileDialog 例子
   
  
'''
import sys
import openpyxl
import pandas as pd  # 引入pandas
import os

from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *


log = print



def txt_to_excel(path):
    # filename = "微相.txt"
    filepath = path

    papa = pd.read_csv(filepath_or_buffer=filepath, sep='\t', encoding="GBK")  # 加载papa.txt,指定它的分隔符是 \t
    # papa.set_index('井名')
    writer = pd.ExcelWriter('微相111output.xlsx')
    papa.to_excel(writer, '微相')
    writer.save()

#
# def load_xlsx(papa="", path=''):
#     # papa是pandas打开的文件
#
#     papa = ''
#
#     filename = '微相.txtname.xlsx'
#     workbook = openpyxl.load_workbook(filename)
#     # 所有的工作表单名字
#     # log(workbook.sheetnames)
#
#     # 拿到某个 工作表 内容的标准做法，我们这个 excel 文件只有一个三条曲线数据工作表
#     sheet = workbook['sheet1']
#
#     maxrow = sheet.max_row  # 最大行数
#     log('maxrow', maxrow)
#
#     log('row', sheet['B'])
#     # # 遍历第二行所有格子
#     # for c in sheet['2']:
#     #     log('格子({})'.format(c.value))
#     # # 遍历获取 A 列所有的值
#     # log('start time')
#
#     log('start time')
#     wells = []
#     wels = set()
#     for a in sheet['B']:
#         # log('a', a.value)
#         if a.value == "井名":
#             head = a.value
#         else:
#             wells.append(a.value)
#         # wels.add(a)
#     # wels = set(wells)
#
#     news_ids = []
#     for id in wells:
#         if id not in news_ids:
#             news_ids.append(id)
#     # wels.clear()
#     log('end time', len(news_ids))
#     log('end time', news_ids)
#
#     # 得到不重复的井名, 保存为excel
#     # path = 'G:/2018研1一/test.txt'
#     for w in news_ids:
#         c = papa[papa['井名'].isin([w])]
#         wname = '{}井.xlsx'.format(w)
#         # log()
#         writer = pd.ExcelWriter(wname)
#
#         # 输出成txt
#         c.to_csv(path_or_buf=path, sep=' ')
#
#         # 输出成excel
#         c.to_excel(writer, 'Sheet1')
#         # df2.to_excel(writer, 'Sheet2')
#         writer.save()
#         # log("我= " , w)
#         break
#


class filedialogdemo(QWidget):
    def __init__(self, parent=None):
        super(filedialogdemo, self).__init__(parent)
        layout_grid = QGridLayout()
        # layout_grid.setHorizontalSpacing(2)
        # layout_grid.setVerticalSpacing(2)

        self.text_load = QTextEdit()
        self.text_load.setReadOnly(True)
        layout_grid.addWidget(self.text_load,0,0)
        btn_load = QPushButton("选择txt路径")
        btn_load.clicked.connect(self.getLoadFileWithPath)
        layout_grid.addWidget(btn_load,0, 1)

        self.text_save = QTextEdit()
        self.text_save.setReadOnly(True)
        layout_grid.addWidget(self.text_save,1, 0)
        btn_save = QPushButton("选择保存位置")
        btn_save.clicked.connect(self.getSaveFileWithPath)
        layout_grid.addWidget(btn_save,1, 1)

        # self.save_excle_btn = QPushButton("保存excel")
        # self.save_excle_btn.clicked.connect(self.save_excel_file)
        # layout.addWidget(self.save_excle_btn)

        btn_process= QPushButton("处理文件")
        btn_process.clicked.connect(self.processFile)
        layout_grid.addWidget(btn_process,2, 1)

        self.setLayout(layout_grid)
        self.setWindowTitle("File Dialog 例子")
        #初始化，以便于后面进行判断
        self.loadedFileName=''
        self.saveDirPath=''

    def getLoadFileWithPath(self):
        log(" 点击getLoadFileWithPath（）")
        dlg = QFileDialog()
        dlg.setFileMode(QFileDialog.AnyFile)
        dlg.setFilter(QDir.Files)
        if dlg.exec_():
            filenames = dlg.selectedFiles()
            self.loadedFileName = filenames[0]
            log("加载的文件filename({})".format(self.loadedFileName))
            self.text_load.setText(self.loadedFileName)
            return filenames[0]

    def getSaveFileWithPath(self):
        # fname, _ = QFileDialog.getSaveFileName(self,
        #                                              "文件保存",
        #                                              "d:/",
        #                                              "Text Files (*.txt)")
        # 这里不是选择文件夹，是选择路径，参考：https: // blog.csdn.net / jirryzhang / article / details / 59088964
        self.saveDirPath = QFileDialog.getExistingDirectory()
        # self.le.setPixmap(QPixmap(fname))
        #
        # # 获取保存的文件夹地址
        # [dirname, filename] = os.path.split(fname)
        # print(dirname, "\n", filename)
        # self.saveDirPath = dir_path
        # self.xlsx_to_txt()
        log('保存文件路径为({})'.format(self.saveDirPath))
        self.text_save.setText(self.saveDirPath)

    # def getLoadFileWithPath(self):
    #     fname, _ = QFileDialog.getOpenFileName(self, 'Open file', 'c:\\', "All Files (*);;Text Files (*.xlsx)")
    #     self.le.setPixmap(QPixmap(fname))

    # def save_excel_file(self):
    # #     fname, _ = QFileDialog.getSaveFileName(self,
    # #                                                  "文件保存",
    # #                                                  "C:/",
    # #                                                  "Text Files (*.xlsx)")
    # #
    # #     self.le.setPixmap(QPixmap(fname))
    # #     # 获取保存的文件夹地址
    # #     [dirname, filename] = os.path.split(fname)
    # #     log('debug save_excel_file 保存文件({})'.format(dirname))
    # #     print(dirname, "\n", filename)
    # #     # 保存文件到excel
    # #     self.save_path_name_excel = dirname
    # #     self.xlsx_to_excel()
    # #     log('debug 保存文件({})'.format(fname))


    def processFile(self):
        log(" 点击processFile（）")
        if ''==self.loadedFileName:
            QMessageBox.warning(self,#使用infomation信息框
                                    "操作错误",
                                    "请选择输入文件！",
                                    QMessageBox.Ok)
            return
        if ''==self.saveDirPath:
            QMessageBox.warning(self,#使用infomation信息框
                                    "操作错误",
                                    "请选择输出文件路径！",
                                    QMessageBox.Ok)
            return
        # https: // www.cnblogs.com / jhao / p / 7243043.html
        if not (os.path.exists(self.loadedFileName)):
            QMessageBox.warning(self,#使用infomation信息框
                                    "操作错误",
                                    "输入文件非法！",
                                    QMessageBox.Ok)
            return
        if not (os.path.exists(self.saveDirPath)):
            QMessageBox.warning(self,#使用infomation信息框
                                    "操作错误",
                                    "输出路径非法！",
                                    QMessageBox.Ok)
            return
        # f = open(self.loadedFileName, 'r')
        # with f:
        #     data = f.read()
        #     self.text_contents.setText(data)
        # log("文件读入完成")
        ifSuccess=self.txt_to_txt(self.loadedFileName,self.saveDirPath)
        if ifSuccess:
            self.msg_process_success()
        #
        # # data = self.getfiles()
        # # log(type(data))
        # # # data_list = data.split('\r\n')
        # # log(type(data), data)
        pass

    def txt_to_txt(self, loadedFileName,dirPath):
        assert os.path.exists(loadedFileName)
        assert os.path.exists(dirPath)
        # filename = "微相.txt"
        # filepath = fpath
        # log("保存文件的文件名字({})".format(filepath))
        # filename = loadedFileName.split('/')[-1]
        # papa = pd.read_csv(filepath_or_buffer=filepath, sep='\t', encoding="GBK")  # 加载papa.txt,指定它的分隔符是 \t
        # # papa.set_index('井名')
        # # log("打开文件({})", papa)
        # self.pdates = papa
        # writer = pd.ExcelWriter('{}.xlsx'.format(filename))
        # self.xlsx_processed_name = '{}.xlsx'.format(filename)
        # papa.to_excel(writer, 'sheet1')
        # self.xlsx_sheet = 'sheet1'
        # log("到ec 保存")
        # writer.save()
        # log("end 保存")
        #对padas不熟悉，就用最简单的，参考： https: // blog.csdn.net / xiaopihaierletian / article / details / 72530682
        f = open(loadedFileName, "r")
        sourceInLine = f.readlines()  # 读取全部内容
        # for line in lines:
        #     lineData = line.strip().split(',')  # 去除空白和逗号“,”
        #     print(lineData)
        # https: // blog.csdn.net / littlle_yan / article / details / 79302488
        # listData=[]
        dictWellData={}#一个井一个字典值，防重复
        iLineNum=0
        for line in sourceInLine:
            iLineNum = iLineNum + 1
            if 1==iLineNum:#第一行是标题，丢掉
                continue
            temp1=line.strip('\n')
            temp2=temp1.strip().split('\t')   # 去除空白、逗号“,”、tab
            if temp2==['']:#防有空格的空行
                log('debug：文件第%d行是空行'.format(iLineNum))
                continue
            if len(temp2)!=3: #3列数据才对
                log('debug：文件第%d行格式有错误'.format(iLineNum))
                continue
            temp2[1]=float(temp2[1])
            temp2[2] = int(temp2[2])
            if temp2[0] in dictWellData.keys():
                dictWellData[temp2[0]].append([temp2[1], temp2[2]])
            else:
                dictWellData[temp2[0]] = []
                dictWellData[temp2[0]].append([temp2[1], temp2[2]])
        print(dictWellData)
        for strWellName in dictWellData.keys():#遍历所有井，参考https://www.cnblogs.com/stuqx/p/7291948.html
            fileName=os.path.join(dirPath, strWellName+'.txt')#https://blog.csdn.net/qq_42034590/article/details/80031241
            f1 = open(fileName, 'w')
            listData=dictWellData[strWellName]
            for i in range(len(listData)):
                # strTemp="%f\t%d\n"%(listData[i][0],listData[i][1])
                strTemp =str(listData[i][0])+"\t"+str(listData[i][1])+"\n"
                f1.write(strTemp)


    def xlsx_to_txt(self, papa=""):
        # papa是pandas打开的文件
        log("开始保存处理过的数据")
        papa = self.pdates

        filename = self.xlsx_processed_name
        log("filename", self.file_loaded_name)
        workbook = openpyxl.load_workbook(filename)
        log("创建了work")
        # 所有的工作表单名字
        # log(workbook.sheetnames)

        # 拿到某个 工作表 内容的标准做法，我们这个 excel 文件只有一个三条曲线数据工作表
        sheet = workbook['sheet1']

        maxrow = sheet.max_row  # 最大行数
        log('maxrow', maxrow)

        log('row', sheet['B'])
        # # 遍历第二行所有格子
        # for c in sheet['2']:
        #     log('格子({})'.format(c.value))
        # # 遍历获取 A 列所有的值
        # log('start time')

        # log('start time')
        wells = []
        wels = set()
        for a in sheet['B']:
            # log('a', a.value)
            if a.value == "井名":
                head = a.value
            else:
                wells.append(a.value)
            # wels.add(a)
        # wels = set(wells)

        news_ids = []
        for id in wells:
            if id not in news_ids:
                news_ids.append(id)
        # wels.clear()
        log('end time', len(news_ids))
        # log('end time', news_ids)

        # 得到不重复的井名, 保存为excel
        path = self.save_path_name

        log("txt path", path)

        for w in news_ids:
            c = papa[papa['井名'].isin([w])]
            # wname = '{}井.xlsx'.format(w)
            wname = '{}井.txt'.format(w)
            log(wname)
            log(path)
            # dparh =
            dpath = path + '/' + wname
            # writer = pd.ExcelWriter(wname)

            log('# 输出成txt', dpath)

            c.to_csv(path_or_buf=dpath, sep=' ')
            log('#完成 输出成txt')

            # 输出成excel
            # c.to_excel(writer, 'Sheet1')
            # df2.to_excel(writer, 'Sheet2')
            # writer.save()
            log("保存txt结束 ", w)
            # break

        # 弹出保存数据成功
        self.msg_save_txt_success()

    def xlsx_to_excel(self, papa=""):
        # papa是pandas打开的文件
        log("开始保存处理过的数据 到 excel")
        papa = self.pdates

        filename = self.xlsx_processed_name
        log("filename", self.xlsx_processed_name)
        workbook = openpyxl.load_workbook(filename)
        log("创建了work")
        # 所有的工作表单名字
        # log(workbook.sheetnames)
        sheet = workbook['sheet1']

        maxrow = sheet.max_row  # 最大行数
        log('maxrow', maxrow)
        wells = []
        wels = set()
        for a in sheet['B']:
            # log('a', a.value)
            if a.value == "井名":
                head = a.value
            else:
                wells.append(a.value)

        # 井名去重
        news_ids = []
        for id in wells:
            if id not in news_ids:
                news_ids.append(id)
        log('当前共口{}井'.format(len(news_ids)))

        # 得到不重复的井名, 保存为excel
        # path为选中保存的dir
        path = self.save_path_name_excel
        # log('end time({}) ({})'.format(len(path), path))

        for w in news_ids:
            c = papa[papa['井名'].isin([w])]
            # wname = '{}井.xlsx'.format(w)
            wname = '{}井.xlsx'.format(w)
            log(wname)
            log(path)
            epath = path

            # writer = pd.ExcelWriter(wname)
            log('# 输出成excel', epath)
            # 输出成excel
            # path = 'C:\'

            writer = pd.ExcelWriter('{}/{}.xlsx'.format(epath, wname)) # 指定Excel文件名与路径
            c.to_excel(writer, sheet_name='Sheet1')
            writer.save()
            # break
        log("保存excel结束 ", w)
        # 弹出保存成功窗口
        self.msg_save_excel_success()


    def msg_process_success(self, msg=''):
        # https: // blog.csdn.net / a359680405 / article / details / 45152131
        # 弹出关于此软件的信息
        msg = """ 文件处理:

        操作处理成功
        """
        # QMessageBox.about(self, "About the demo", msg.strip())
        QMessageBox.information (self, "数据处理成功", msg.strip())
    #
    # def msg_save_txt_success(self, msg=''):
    #     # 弹出关于此软件的信息
    #     msg = """ 保存成txt文件:
    #
    #     操作处理成功
    #     """
    #     # QMessageBox.about(self, "About the demo", msg.strip())
    #     QMessageBox.information (self, "数据处理成功", msg.strip())
    #
    # def msg_save_excel_success(self, msg=''):
    #     # 弹出关于此软件的信息
    #     msg = """ 保存成excel文件:
    #
    #      操作处理成功
    #      """
    #     # QMessageBox.about(self, "About the demo", msg.strip())
    #     QMessageBox.information (self, "数据处理成功", msg.strip())

def test():
    # app = QApplication(sys.argv)
    # ex = filedialogdemo()
    # ex.process_file()

    txt_to_excel("微相.txt")
    log("测试完成")

# 1.加载excel数据
# 2.根据excel数据切割



if __name__ == '__main__':
    # # https: // blog.csdn.net / Avada_533 / article / details / 78762773
    # Back up the reference to the exceptionhook
    sys._excepthook = sys.excepthook
    def my_exception_hook(exctype, value, traceback):
        # Print the error and traceback
        print(exctype, value, traceback)
        # Call the normal Exception hook after
        sys._excepthook(exctype, value, traceback)
        sys.exit(1)
    # Set the exception hook to our wrapping function
    sys.excepthook = my_exception_hook

    app = QApplication(sys.argv)
    ex = filedialogdemo()
    ex.show()

    sys.exit(app.exec_())
    try:
        sys.exit(app.exec_())
    except:
        print("Exiting")
