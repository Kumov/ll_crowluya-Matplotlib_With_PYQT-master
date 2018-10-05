import pandas as pd #引入pandas
import openpyxl
from utils import log
filename = "微相.txt"
papa = pd.read_csv(filename, sep='\t', encoding="GBK")  # 加载papa.txt,指定它的分隔符是 \t

def txt_to_excel():
    filename = "微相.txt"
    papa = pd.read_csv(filename, sep='\t',encoding="GBK") #加载papa.txt,指定它的分隔符是 \t
    papa.set_index('井名')
    writer = pd.ExcelWriter('微相111output.xlsx')
    papa.to_excel(writer, '微相')
    # df2.to_excel(writer,'Sheet2')
    writer.save()


def load_xlsx():
    filename = '微相.txtname.xlsx'
    workbook = openpyxl.load_workbook(filename)
    # 所有的工作表单名字
    # log(workbook.sheetnames)

    # 拿到某个 工作表 内容的标准做法，我们这个 excel 文件只有一个三条曲线数据工作表
    sheet = workbook['sheet1']

    maxrow = sheet.max_row         # 最大行数
    log('maxrow',  maxrow)

    log('row', sheet['B'])
    # # 遍历第二行所有格子
    # for c in sheet['2']:
    #     log('格子({})'.format(c.value))
    # # 遍历获取 A 列所有的值
    # log('start time')

    log('start time')
    wells = []
    wels = set()
    for a in sheet['B']:
        # log('a', a.value)
        if a.value == "井名":
            head = a.value
        else:
            wells.append(a.value)
        # wels.add(a)
    # wels = set(wells)

    news_ids = []
    for id in wells:
        if id not in news_ids:
            news_ids.append(id)
    # wels.clear()
    log('end time', len(news_ids))
    log('end time', news_ids)

    # 得到不重复的井名, 保存为excel
    path = 'G:/2018研1一/test.txt'
    for w in news_ids:
        c = papa[papa['井名'].isin([w])]
        wname = '{}井.xlsx'.format(w)
        # log()
        writer = pd.ExcelWriter(wname)

        # 输出成txt
        # c.to_csv(path_or_buf = path, sep=' ')

        # 输出成excel
        c.to_excel(writer, 'Sheet1')
        # df2.to_excel(writer, 'Sheet2')
        writer.save()
        # log("我= " , w)
        break
    # 获得重复井的位置

    # 切割重复的井数据






def test():
    # log(dataframe.井名.duplicated())
    load_xlsx()
    # papa.set_index('井名')
    # log(papa.axes)
    log(papa.axes)
    w = 'H38-61'
    # log(papa[papa['井名'].isin(w)])
    log(papa[papa['井名'].isin([w])])
    # papa[papa['井名'].isin(['H38-61'])]
    # papa[papa['井名'].isin(['H38-61'])]

    pass

if __name__ == '__main__':
    # txt_to_excel()
    # dataframe = papa
    test()
