import openpyxl
import time

'''
本项目需要使用 openpyxl 处理 excel 文件
注意，只支持 xlsx 格式，不支持 xls 格式

需要用 pip 安装 openpyxl 库

本项目和项目 11 是一样的，不同之处在于本次要处理和生成的是 excel 文件而不是 csv 文件
参考下面的例子代码，实现对 12-list.xlsx 的处理，这个文件在群文件中
处理要求同项目 11 一样


openpyxl 的官方文档（本次项目没有看的必要）:
http://openpyxl.readthedocs.io/en/stable/
'''


def time_log():
    value = time.localtime(int(time.time()))
    return value


def log(*args, **kwargs):
    # time.time() 返回 unix time
    # 如何把 unix time 转换为普通人类可以看懂的格式呢？
    format = '%Y/%m/%d %H:%M:%S'
    value = time.localtime(int(time.time()))
    dt = time.strftime(format, value)
    print(dt, *args, **kwargs)


def write(t):
    # 创建一个新的工作表簿
    workbook = openpyxl.Workbook()
    # 获取表格的默认工作表
    sheet = workbook.active
    # 设置工作表的凑题
    sheet.title = '新标题'
    # 在 C栏 的第一行写入数据
    sheet['d1'] = 'hello python'
    # 在 A栏 中写入数据
    log('start ({})times'.format(t))
    for i in range(t):
        s = 'A{}'.format(i + 1)
        s1 = 'B{}'.format(i + 1)
        s2 = 'C{}'.format(i + 1)

        sheet[s].value = i + 1
        sheet[s1].value = i + 1
        sheet[s2].value = i + 1

    # 保存到 demo.xlsx 文件中
    log('end write({}) times'.format(t))
    workbook.save('demo 次数 = ({}).xlsx'.format(t))
    log('end save({}) t')


def read(f=[], sheet_number=0):
    # filename = 'demo 次数 = (1000000).xlsx'
    f = 'demosave.xlsx'
    filename = f
    workbook = openpyxl.load_workbook(filename)
    # 所有的工作表单名字
    log(workbook.sheetnames)
    # nowsheet ='Sheet1'
    sheet = workbook.sheetnames[sheet_number]
    # 拿到某个 工作表 内容的标准做法，我们这个 excel 文件只有一个 city 工作表
    sheet = workbook[sheet]
    # sheet['C1'].value     # 返回 C 列第 1 行格子里的值
    # sheet['1']            # 第一行的所有格子(返回一个 tuple)
    # sheet.max_column      # 最大列数
    # sheet.max_row         # 最大行数
    header = []
    body = []
    excel_data = [header, body]
    log(sheet.rows)
    # 注意range从1开始的，因为在openpyxl中为了和Excel中的表达方式一致，并不和编程语言的习惯以0表示第一个值。
    # log('sheet.rows({}) sheet.rows[0] ({})'.format(sheet.rows, sheet.rows[0]))

    for cell in list(sheet.rows)[2]:
        print(cell.value, type(cell.value))
    log("sheet values({})".format(sheet.values))
    for i in range(sheet.max_row):
        # 第一行为标题
        row_data = []
        # row_raw_data = sheet[i+1]
        for c in sheet[i+1]:
            # log('格子({})'.format(c.value))
            row_data.append(c.value)
        log('rowdata ({})'.format(row_data))

        # row_data = sheet[str(i+1)]
        #
        if i == 0:
            excel_data[0] = row_data
        else:
            body.append(row_data)
        log('header ({}) body ({})'.format(header, body))
        pass

    log('excel data = ({})'.format(excel_data))
    # # 用 .value 获取格子里的值
    # log('sheet c1 value', sheet['C1'].value)
    # log('max row', sheet.max_row)
    # # # 获取一行
    # log('row', sheet['2'])
    # # # # 遍历第二行所有格子
    # for c in sheet['2']:
    #     log('格子({})'.format(c.value))
    # # # # 遍历获取 A 列所有的值
    # log('start time')
    # for a in sheet['A']:
    #     log('a', a.value)
    # log('end time')


def test_write():
    # 传入不同的写入次数 + 保存文件 测试消耗的时间
    # 1w 10w 100w
    # 1s  7s  111s

    t = [10000, 100000, 1000000]
    for item in t:
        write(item)
    pass


def test_read():
    # 传入不同的读入次数 测试消耗的时间
    # 1w 10w 100w
    # 1s  10s  111s
    t = ['demo 次数 = (10000).xlsx', 'demo 次数 = (100000).xlsx', 'demo 次数 = (1000000).xlsx']

    read(t[1])

    # for item in t:
    #     read(item)
    # pass


def test():
    # test_write()
    # test_read()
    read()
    pass


def main():
    # read()
    write()
    pass


if __name__ == '__main__':
    # main()
    test()
