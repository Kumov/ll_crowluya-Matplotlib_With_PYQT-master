#
import pandas as pd
import numpy as np
import openpyxl
import xlrd
import xlwt

import os
# read_excel
# to_excel
log = print

# https://blog.csdn.net/zhxpingmiss/article/details/78705004  # 1.读入xls 或者xlsx文件(选中文件路径) python操作xls、xlsx格式Excel方法： .
# https://blog.csdn.net/liumang9438/article/details/80101863   Python中 win32ui 模块打开文件和另存为对话框的简单实现
# http://timgolden.me.uk/python/win32_how_do_i/browse-for-a-folder.html 浏览选中文件夹
loadedFileName = ''
saveDirPath = ''
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *


def end_with_allowed_end(file, allowed_end=['.xls', '.xlsx']):
    # 获得指定文件后缀
    # 用于出错处理 但是没有到
    filname = file
    end = os.path.splitext(filname)[1]
    log("end", end)
    # acc
    # allowed_end = '.txt'
    if end not in allowed_end:
        log('文件名字', filname)
        # self.loadedFileName.split()
        # QMessageBox.warning(self,  # 使用infomation信息框
        #                     "操作错误",
        #                     "请选择txt格式文件！",
        #                     QMessageBox.Ok)
        # self.text_load.setText("")
        return False
    else:
        return True


def load_excel_file():
    # 弹窗获得excel文件位置
    import win32ui
    dlg = win32ui.CreateFileDialog(1)  # 1表示打开文件对话框
    # dlg.SetOFNInitialDir('E:/Python')  # 设置打开文件对话框中的初始显示目录
    dlg.DoModal()
    filename = dlg.GetPathName()  # 获取选择的文件名称
    # dlg.GetFileName()
    # dlg.GetFileTitle()
    # dlg.GetPathName()
    # dlg.GetPathNames()
    # help(dlg)
    # print(filename)
    # 看是否是自己需要的文件
    # end_with_allowed_end(filename)
    # log(filename)

    return filename
#
#
# def load_excel():
#     # 加载用户的文件
#     pass


def read_xlsx(f=None, sheet_number=0):
    import openpyxl
    # 注意range从1开始的，因为在openpyxl中为了和Excel中的表达方式一致，并不和编程语言的习惯以0表示第一个值。
    # filename = 'demo 次数 = (1000000).xlsx'
    # f = 'demosave.xlsx'

    if f is None:
        return False
    filename = f
    workbook = openpyxl.load_workbook(filename)
    # 所有的工作表单名字
    log(workbook.sheetnames)
    # 获取当前工作表 默认为第一个
    sheet_name = workbook.sheetnames[sheet_number]
    sheet = workbook[sheet_name]

    header = []
    body = []
    excel_data = [header, body]
    # log(sheet.rows)
    # 注意range从1开始的，因为在openpyxl中为了和Excel中的表达方式一致，并不和编程语言的习惯以0表示第一个值。
    # log('sheet.rows({}) sheet.rows[0] ({})'.format(sheet.rows, sheet.rows[0]))

    for i in range(sheet.max_row):
        # 第一行为标题
        row_data = []
        # row_raw_data = sheet[i+1]
        # 注意range从1开始的，因为在openpyxl中为了和Excel中的表达方式一致，并不和编程语言的习惯以0表示第一个值。
        # 要获取单元格的值只有一个单元一个单元的获取
        for c in sheet[i + 1]:
            row_data.append(c.value)
        if i == 0:
            excel_data[0] = row_data
        else:
            body.append(row_data)
        # log('header ({}) body ({})'.format(header, body))
        pass

    return excel_data


def read_xls(f, sheet_number=0):
    filename = f
    sheet_num = sheet_number
    # filename = "生产数据.xls"
    data = xlrd.open_workbook(filename)
    # 所有表名字
    sheetNames = data.sheet_names()
    # 默认打开第一个表的数据
    header = []
    body = []
    excel_data = [header, body]
    oneSheetData = data.sheets()[sheet_num]
    # 获取当前表的行数
    crowlength = oneSheetData.nrows

    # 第一行
    # #第 j 行数据
    for i in range(crowlength):
        log('i {}  '.format(i))
        rowData = oneSheetData.row_values(i)
        if i == 0:
            excel_data[0] = rowData
            # log('len header ({})'.format(header))
        else:
            body.append(rowData)
    log('excel data', excel_data)

    return excel_data
    pass


def read_excel(file, sheet_number=0):
    # 读取excel文件 不同后缀调用不同方法
    # 如果文件名xls结尾

    # 1获取当前文件的文件后缀名
    filename = file
    end = os.path.splitext(filename)[1]
    log('file end', end)
    exceldata = None
    # end_with_xls = ''
    # end_with_xlsx = ''

    if end == '.xls':
        exceldata = read_xls(filename, sheet_number)
    # 如果文件名xlsx结尾
    elif end == '.xlsx':
        exceldata = read_xlsx(filename, sheet_number)

    return exceldata


def write_excel(data=[], filename=""):
    # 获取list 文件名 保存
    log('开始写文档里', filename)
    from openpyxl import Workbook
    # 保存为excel
    # 打开工作空间
    # datalist = read_xls(0)
    datalist = data
    # filname = file

    # 没有数据 返回false
    if len(data) == 0:
        return False
    # 如果文件有后缀
    # 如果文件没有后缀

    end = os.path.splitext(filename)[1]
    allowed_end = ['.xls', '.xlsx']
    if end is None:
        log(' end i s none')
        pass
    else:

        pass
    log("文件结尾 end", end, type(end), len(end))
    fileName = filename + '.xlsx'
    log("保存之后的文件名 ({})".format(fileName))
    # return
    # fileName = 'demosave' + '.xlsx'
    wb = Workbook()
    ws = wb.active
    # 保存当前井数据到excel
    #
    header = datalist[0]
    listData = datalist[1]
    # log('header({})'.format(header))
    # log('listData({})'.format(listData))

    ws.append(header)
    for row in listData:
        ws.append(row)
    wb.save(fileName)
    log('文档写完了')
    pass


def get_save_path():
    # 获取保存的文件夹
    import win32gui
    from win32com.shell import shell, shellcon

    desktop_pidl = shell.SHGetFolderLocation(0, shellcon.CSIDL_DESKTOP, 0, 0)
    pidl, display_name, image_list = shell.SHBrowseForFolder(
        win32gui.GetDesktopWindow(),
        desktop_pidl,
        "Choose a folder",
        0,
        None,
        None
    )
    # 获取
    print(shell.SHGetPathFromIDList(pidl))
    l = shell.SHGetPathFromIDList(pidl)
    print(l)
    return l
    # 获得输入的文件名

    # return filename

    pass


def get_file_saved_path():
    # 获取保存的文件及路径
    import win32ui

    dlg = win32ui.CreateFileDialog(0)
    # 默认为打开桌面
    dlg.SetOFNInitialDir("C:/Users/Administrator/Desktop")
    flag = dlg.DoModal()
    log(flag)
    log(dlg.GetPathName())
    if 1 == flag:
        log(dlg.GetPathName())
        log('文件地址 （{}）'.format(dlg.GetPathName()))
        return dlg.GetPathName()
    else:
        log("取消另存为...")
        return False


def test():
    # read_xls(0)
    # write_excel()
    # getLoadFileWithPath()
    # load_excel_file()
    # get_save_path()
    get_file_saved_path()
    pass


def main():
    import time
    file = load_excel_file()
    log(" file = ({})".format(file))
    dataList = read_excel(file)
    log(" dataList = ({})".format(dataList))
    # 获取另存为的地址
    for i in range(3):
        log('处理中, 3秒钟之后保存, 保存位置随选， 文件名随选', i)
        time.sleep(1)
    saved_file_path = get_file_saved_path()
    log(" 文件保存路径 = ({})".format(saved_file_path))
    # 保存文件
    write_excel(data=dataList, filename=saved_file_path)
    pass


if __name__ == '__main__':
    # test()
    main()
    pass






