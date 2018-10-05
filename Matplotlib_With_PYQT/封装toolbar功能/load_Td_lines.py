# 加载excel文件
# 保存到txt
# 保存到excel
import json
import openpyxl
import time
from models.td import Td
import matplotlib.pyplot as plt  # 导入绘图包
import matplotlib as mpl
import numpy as np

mpl.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
mpl.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号


# class TD(object):
#     def __init__(self):
#
#         pass

# 全局数据在这里
# tDs = []
def save(data, path):
    """
    data 是 dict 或者 list
    path 是保存文件的路径
    """
    tDs = data

    # classname = cls.__name__
    filename = 'td'
    path = 'data/{}.txt'.format(filename)
    log(' path = ({})'.format(path))
    with open(path, 'w+', encoding='utf-8') as f:
        # s = f.read()
        for i in range(len(tDs)):
            # strTemp="%f\t%d\n"%(listData[i][0],listData[i][1])

            # strTemp = str(listData[i][0]) + "\t" + str(listData[i][1]) + "\n"
            for t in tDs[i]:
                log("t", t)
                f.write(str(t) + '\n')


def load(path):
    with open(path, 'r', encoding='utf-8') as f:
        s = f.read()


def log(*args, **kwargs):
    '''
     log为自己写的工具类 打断点 方便调试
    :param args:
    :param kwargs:
    :return:
    '''
    # time.time() 返回 unix time
    # 如何把 unix time 转换为普通人类可以看懂的格式
    format = '%Y/%m/%d %H:%M:%S'
    value = time.localtime(int(time.time()))
    dt = time.strftime(format, value)
    print(dt, *args, **kwargs)


def load_xlsx():
    # 接收一个.xlsx文件
    # 以json格式存储在data/td.txt文件中
    # 本文件可以以字典的形式存储
    # td
    # list_rD=:
    # list_TD=:
    filename = '三条tD曲线数据.xlsx'
    workbook = openpyxl.load_workbook(filename)
    # 所有的工作表单名字
    # log(workbook.sheetnames)

    # 拿到某个 工作表 内容的标准做法，我们这个 excel 文件只有一个三条曲线数据工作表
    sheet = workbook['三条曲线数据']
    # sheet['C1'].value     # 返回 C 列第 1 行格子里的值
    # sheet['1']            # 第一行的所有格子(返回一个 tuple)
    # sheet.max_column      # 最大列数
    # sheet.max_row         # 最大行数

    #  获取不同A 列 td的值

    # 用 .value 获取格子里的值
    # log('sheet A1 value', sheet['C1'].value)
    # log('max row', sheet.max_row)
    # 获取一行
    # log('row', sheet['2'])
    # # 遍历第二行所有格子
    # for c in sheet['2']:
    #     log('格子({})'.format(c.value))
    # # 遍历获取 A 列所有的值
    # log('start time')

    le = len(sheet['A'])
    log("debug le ({})".format(le, sheet['A']))
    i = 0
    td_cell = []
    for a in sheet['A']:
        # td_cell = []
        v = a.value
        log('数据类型 ({})'.format(type(v)))
        td_cell.append(v)
        i = i + 1

    # td_cell保存了xlsx文件所有信息
    # 数据全是 str 保存的

    log('end time cell ({})'.format(td_cell))

    # 用类来保存
    # 每隔六行 新建一个 td类
    dis = 6

    tDs =[]
    for i in range(0, le, dis):
        cell = []
        line_rd = []
        line_TD = []

        # title = td_cell[i].split('=')[1]
        title = td_cell[i]

        line_rd.append(td_cell[i + 1])
        line_rd.append(td_cell[i + 2])

        line_TD.append(td_cell[i + 3])
        line_TD.append(td_cell[i + 4])

        # 单挑数据组装成 一个cell
        cell.append(title)
        cell.append(line_rd)
        cell.append(line_TD)

        log('cells ', cell)
        # 所有曲线数据
        tDs.append(cell)
        # rd_datas = td_cell[i + 2]
        # TD_datas = td_cell[i + 4]

    # 将数据保存到txt中



def test():
    load_xlsx()
    # log('tds')
    pass


if __name__ == '__main__':
    test()
